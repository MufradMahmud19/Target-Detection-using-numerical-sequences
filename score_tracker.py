"""
Score Tracker Module
Manages an Excel workbook (scores.xlsx) that records run results.
Structure:
  - One sheet per unique target coordinate (e.g., "Target_200_300")
  - Each sheet has rows for each series type used at that target
  - Tracks: Best Time, Latest Time, Attempt Count, Last Updated
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

SCORES_FILE = "scores.xlsx"


def _get_scores_path():
    """Get the absolute path to the scores Excel file."""
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), SCORES_FILE)


def _get_sheet_name(target_x, target_y):
    """Generate a sheet name from target coordinates."""
    return "Target_{}_{}".format(int(target_x), int(target_y))


def _style_header_row(ws):
    """Apply styling to the header row of a worksheet."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B547E", end_color="2B547E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="4A7FB5")
    )

    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 22


def is_available():
    """Check if score tracking is available (openpyxl installed)."""
    return HAS_OPENPYXL


def update_score(target_x, target_y, series_name, time_seconds):
    """
    Record a run result in the Excel workbook.

    Args:
        target_x: Target X coordinate
        target_y: Target Y coordinate
        series_name: Name of the series used (e.g., "Fibonacci Series")
        time_seconds: Time taken to reach the target in seconds

    Returns:
        The best time for this target+series combination (after update),
        or None if tracking is unavailable.
    """
    if not HAS_OPENPYXL:
        print("[Score Tracker] openpyxl not installed. Run: pip install openpyxl")
        return None

    filepath = _get_scores_path()

    try:
        # Open or create workbook
        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath)
        else:
            wb = openpyxl.Workbook()
            # Remove the default "Sheet" sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        sheet_name = _get_sheet_name(target_x, target_y)

        # Create sheet if this is a new target coordinate
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            headers = ["Series", "Best Time (s)", "Latest Time (s)", "Attempts", "Last Updated"]
            ws.append(headers)
            _style_header_row(ws)
            print("[Score Tracker] Created new sheet: {}".format(sheet_name))
        else:
            ws = wb[sheet_name]

        # Find existing row for this series
        series_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == series_name:
                series_row = row
                break

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rounded_time = round(time_seconds, 3)

        if series_row:
            # Update existing entry
            current_best = ws.cell(row=series_row, column=2).value
            attempts = ws.cell(row=series_row, column=4).value or 0

            # Update best time only if new time is better (lower)
            if current_best is None or rounded_time < current_best:
                best_time = rounded_time
            else:
                best_time = current_best

            ws.cell(row=series_row, column=2, value=best_time)
            ws.cell(row=series_row, column=3, value=rounded_time)
            ws.cell(row=series_row, column=4, value=attempts + 1)
            ws.cell(row=series_row, column=5, value=now)
        else:
            # New series entry for this target
            ws.append([series_name, rounded_time, rounded_time, 1, now])
            series_row = ws.max_row
            best_time = rounded_time

            # Style the new data row
            data_align = Alignment(horizontal="center")
            for col in range(1, 6):
                ws.cell(row=series_row, column=col).alignment = data_align

        wb.save(filepath)
        print("[Score Tracker] Saved to {} | Sheet: {} | {}: {:.3f}s".format(
            SCORES_FILE, sheet_name, series_name, rounded_time))

        return best_time

    except PermissionError:
        print("[Score Tracker] Cannot write to {} - file may be open in Excel.".format(SCORES_FILE))
        return None
    except Exception as e:
        print("[Score Tracker] Error saving score: {}".format(e))
        return None


def get_best_score(target_x, target_y, series_name):
    """
    Retrieve the best time for a specific target + series combination.

    Returns:
        The best time in seconds, or None if no record exists.
    """
    if not HAS_OPENPYXL:
        return None

    filepath = _get_scores_path()
    if not os.path.exists(filepath):
        return None

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        sheet_name = _get_sheet_name(target_x, target_y)

        if sheet_name not in wb.sheetnames:
            wb.close()
            return None

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == series_name:
                best = row[1].value
                wb.close()
                return best

        wb.close()
        return None

    except Exception as e:
        print("[Score Tracker] Error reading score: {}".format(e))
        return None


def get_all_scores_for_target(target_x, target_y):
    """
    Retrieve all series scores for a specific target coordinate.

    Returns:
        List of dicts: [{"series": str, "best": float, "latest": float, "attempts": int}, ...]
        or empty list if no records.
    """
    if not HAS_OPENPYXL:
        return []

    filepath = _get_scores_path()
    if not os.path.exists(filepath):
        return []

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        sheet_name = _get_sheet_name(target_x, target_y)

        if sheet_name not in wb.sheetnames:
            wb.close()
            return []

        ws = wb[sheet_name]
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                results.append({
                    "series": row[0],
                    "best": row[1],
                    "latest": row[2],
                    "attempts": row[3],
                })

        wb.close()
        return results

    except Exception:
        return []
